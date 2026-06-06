"""
=== PIPELINE PROVENANCE — v28 base (analogue of v3, smaller LLM) ===
CohortX Task 1, pure NXML, NO external sources (no CT.gov / AACT / NCT).
This is an exact clone of pipeline_v3.py EXCEPT eligibility/ages are produced by
llm_extractor_rag_v28 (Mistral-7B-Instruct-v0.3 Q4) instead of the 24B model.
The 24B failed the >=16GB-RAM hardware rule; the 7B fits it.

PER-FIELD METHOD (all CPU-capable):
  - conditions : vocab lookup on title+abstract, LLM fallback (llm_extractor)
  - study_type : rule-based (extractors.extract_study_type)
  - sex        : rule-based (extractors.extract_sex)
  - eligibility, minimum_age, maximum_age : RAG few-shot Mistral-7B
                 (llm_extractor_rag_v28), retrieval via PubMedBERT over train_index.json

OUTPUT: submission_v28_base.csv  (the v7 conditions merge + v17 BART eligibility
        rewrite are applied afterward by build_v28.py to produce submission_v28.csv)

REPRODUCE:
  # pure CPU (competition mode):
  python pipeline_v28.py
  # faster dev run with GPU offload (identical output):
  N_GPU_LAYERS=33 python pipeline_v28.py
  # sanity-check accuracy on the train set:
  python pipeline_v28.py --validate --n 50
"""

import csv
import json
import os
import re
import sys
import ast

# Ensure CUDA 12.1 libs are found at runtime (needed for llama_cpp GPU support)
_cuda_lib = '/usr/local/cuda-12.1/targets/x86_64-linux/lib'
os.environ['LD_LIBRARY_PATH'] = _cuda_lib + ':' + os.environ.get('LD_LIBRARY_PATH', '')

TASK_XLSX = '/home/kkolpetinou/cohort-x-task-1/Task_1.xlsx'
OUTPUT_CSV = '/home/kkolpetinou/cohort-x-task-1/submission_v28_base.csv'
NCT_CACHE_FILE = os.path.join(os.path.dirname(__file__), 'nct_cache.json')

_nct_cache = None

def _get_nct_cache():
    global _nct_cache
    if _nct_cache is None:
        if os.path.exists(NCT_CACHE_FILE):
            with open(NCT_CACHE_FILE) as f:
                _nct_cache = json.load(f)
        else:
            _nct_cache = {}
    return _nct_cache

FIELDNAMES = [
    'pmcids', 'conditions', 'study_type', 'sex',
    'minimum_age', 'maximum_age', 'eligibility_criteria',
]


# ── load extractors ──────────────────────────────────────────────────────────

def _load_extractors():
    sys.path.insert(0, os.path.dirname(__file__))
    from nxml_parser import parse_nxml
    from extractors import extract_sex, extract_study_type, extract_conditions, _vocab_lookup
    from llm_extractor import extract_conditions_with_llm
    from llm_extractor_rag_v28 import extract_eligibility_with_llm  # v28: Mistral-7B
    return parse_nxml, extract_sex, extract_study_type, extract_conditions_with_llm, extract_eligibility_with_llm, _vocab_lookup


# ── read xlsx ────────────────────────────────────────────────────────────────

def read_test_pmcids():
    """Return list of PMC ID strings from the Test sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(TASK_XLSX)
    ws = wb['Test']
    rows = list(ws.iter_rows(values_only=True))
    # First row is header; col 0 is pmcid
    pmcids = []
    for row in rows[1:]:
        if row[0]:
            # IDs are stored as floats in xlsx (e.g. 12279614.0) — convert to int string
            pmcids.append(str(int(float(row[0]))).strip())
    return pmcids


def read_train_rows():
    """Return list of (pmcid, gt_dict) from the Train sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(TASK_XLSX)
    ws = wb['Train']
    rows = list(ws.iter_rows(values_only=True))
    # header: pmcids, conditions, study_type, sex, minimum_age, maximum_age, eligibility_criteria
    header = [str(c).strip() if c else '' for c in rows[0]]
    result = []
    for row in rows[1:]:
        if not row[0]:
            continue
        d = {header[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(header))}
        # Normalise float PMC IDs to plain integers
        d['pmcids'] = str(int(float(d['pmcids'])))
        result.append(d)
    return result


# ── process one article ──────────────────────────────────────────────────────

def _hybrid_conditions(parsed, extract_conditions_with_llm, _vocab_lookup):
    """Use vocab lookup on title+abstract for high-precision GT-name matches.
    Fall back to LLM only when vocab finds nothing (avoids body-text noise)."""
    search_text = parsed.get('title', '') + ' ' + parsed.get('abstract_text', '')
    vocab_results = _vocab_lookup(search_text)

    if vocab_results:
        return str(vocab_results[:8])

    return extract_conditions_with_llm(parsed)


def _normalise_ct_age(val):
    """Convert ClinicalTrials age strings like '18 Years', '6 Months' to our format."""
    if not val:
        return ''
    val = val.strip()
    if re.match(r'^\d+\s+Years?$', val, re.IGNORECASE):
        m = re.match(r'^(\d+)\s+Years?$', val, re.IGNORECASE)
        return f"{m.group(1)} Years"
    # Months — convert to years if >= 12, else keep as-is with "Months"
    m = re.match(r'^(\d+)\s+Months?$', val, re.IGNORECASE)
    if m:
        months = int(m.group(1))
        if months >= 12:
            return f"{months // 12} Years"
        return val  # e.g. "6 Months"
    return val


def process(pmcid, parse_nxml, extract_sex, extract_study_type, extract_conditions_with_llm, extract_eligibility_with_llm, _vocab_lookup):
    """Parse and extract all fields for one PMC ID. Returns a row dict."""
    parsed = parse_nxml(pmcid)
    if parsed is None:
        return {
            'pmcids': pmcid,
            'conditions': str(['Not Specified']),
            'study_type': 'INTERVENTIONAL',
            'sex': 'ALL',
            'minimum_age': '18 Years',
            'maximum_age': 'Not Specified',
            'eligibility_criteria': 'Not Specified',
        }

    # ── conditions: vocab lookup on title+abstract, LLM fallback ──
    conditions = _hybrid_conditions(parsed, extract_conditions_with_llm, _vocab_lookup)

    # ── study_type: rule-based classifier ──
    study_type = extract_study_type(parsed)

    # ── sex: rule-based ──
    sex = extract_sex(parsed)

    # ── eligibility + ages: RAG-augmented Mistral-24B from NXML ──
    llm_result = extract_eligibility_with_llm(parsed)
    eligibility = llm_result['eligibility_criteria'] or 'Not Specified'
    min_age = llm_result['minimum_age'] or 'Not Specified'
    max_age = llm_result['maximum_age'] or 'Not Specified'

    return {
        'pmcids': pmcid,
        'conditions': conditions,
        'study_type': study_type,
        'sex': sex,
        'minimum_age': min_age,
        'maximum_age': max_age,
        'eligibility_criteria': eligibility,
    }


# ── validation helpers ───────────────────────────────────────────────────────

def _norm_conditions(val):
    """Parse a conditions string into a set of lowercase canonical names."""
    if not val or val in ('Not Specified', '[]', ''):
        return set()
    try:
        items = ast.literal_eval(val)
        if isinstance(items, list):
            return {str(x).strip().lower() for x in items}
    except Exception:
        pass
    # Fallback: comma-separated
    return {x.strip().lower() for x in val.split(',') if x.strip()}


def _f1_conditions(pred_str, gt_str):
    pred = _norm_conditions(pred_str)
    gt = _norm_conditions(gt_str)
    if not gt and not pred:
        return 1.0, 1.0, 1.0
    if not gt or not pred:
        return 0.0, 0.0, 0.0
    tp = len(pred & gt)
    p = tp / len(pred) if pred else 0.0
    r = tp / len(gt) if gt else 0.0
    f1 = 2 * p * r / (p + r) if (p + r) > 0 else 0.0
    return p, r, f1


def validate(n_limit=None):
    """Run on Train set and print per-field accuracy / F1."""
    fns = _load_extractors()
    parse_nxml, extract_sex, extract_study_type, extract_conditions_with_llm, extract_eligibility_with_llm, _vocab_lookup = fns

    train_rows = read_train_rows()
    if n_limit:
        train_rows = train_rows[:n_limit]

    total = 0
    correct_type = 0
    correct_sex = 0
    correct_min = 0
    correct_max = 0
    sum_f1_cond = 0.0
    sum_p_cond = 0.0
    sum_r_cond = 0.0
    parse_fails = 0

    for gt in train_rows:
        pmcid = gt['pmcids']
        row = process(pmcid, *fns)
        total += 1

        if parse_nxml(pmcid) is None:
            parse_fails += 1

        if row['study_type'] == gt.get('study_type', '').strip():
            correct_type += 1
        if row['sex'] == gt.get('sex', '').strip():
            correct_sex += 1

        # Ages: GT may be "18 Years", "Not Specified", etc.
        gt_min = gt.get('minimum_age', '').strip()
        gt_max = gt.get('maximum_age', '').strip()
        if row['minimum_age'] == gt_min:
            correct_min += 1
        if row['maximum_age'] == gt_max:
            correct_max += 1

        # Conditions F1
        _, _, f1 = _f1_conditions(row['conditions'], gt.get('conditions', ''))
        p, r, _ = _f1_conditions(row['conditions'], gt.get('conditions', ''))
        sum_f1_cond += f1
        sum_p_cond += p
        sum_r_cond += r

        if total % 50 == 0:
            print(f"  [{total}/{len(train_rows)}] "
                  f"type={correct_type/total:.1%}  "
                  f"sex={correct_sex/total:.1%}  "
                  f"min={correct_min/total:.1%}  "
                  f"max={correct_max/total:.1%}  "
                  f"cond_f1={sum_f1_cond/total:.1%}")

    print(f"\n{'='*60}")
    print(f"Validation over {total} train rows ({parse_fails} parse failures)")
    print(f"  study_type accuracy : {correct_type/total:.1%} ({correct_type}/{total})")
    print(f"  sex accuracy        : {correct_sex/total:.1%} ({correct_sex}/{total})")
    print(f"  minimum_age accuracy: {correct_min/total:.1%} ({correct_min}/{total})")
    print(f"  maximum_age accuracy: {correct_max/total:.1%} ({correct_max}/{total})")
    print(f"  conditions  Precision: {sum_p_cond/total:.1%}")
    print(f"  conditions  Recall   : {sum_r_cond/total:.1%}")
    print(f"  conditions  F1       : {sum_f1_cond/total:.1%}")


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    do_validate = '--validate' in sys.argv
    n_limit = None
    if '--n' in sys.argv:
        idx = sys.argv.index('--n')
        if idx + 1 < len(sys.argv):
            n_limit = int(sys.argv[idx + 1])

    if do_validate:
        print("Running validation on Train set...")
        validate(n_limit=n_limit)
        return

    fns = _load_extractors()
    pmcids = read_test_pmcids()
    print(f"Processing {len(pmcids)} test articles → {OUTPUT_CSV}")

    rows = []
    for i, pmcid in enumerate(pmcids, 1):
        print(f"  [{i}/{len(pmcids)}] PMC{pmcid}...", flush=True)
        row = process(pmcid, *fns)
        rows.append(row)
        print(f"  [{i}/{len(pmcids)}] PMC{pmcid} done", flush=True)

    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Done. Wrote {len(rows)} rows to {OUTPUT_CSV}")


if __name__ == '__main__':
    main()
