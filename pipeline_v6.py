"""
pipeline_v6.py — NXML-only extractor, no LLM.

Lessons learned from train-set comparison:
  * rule-based extract_eligibility BEATS retrieval, verbatim, concat-topk on all 3 proxies
    (FM3S 0.145, token-jaccard 0.138, sequence-ratio 0.071)
  * conditions vocab leaks generics; need stronger blocklist + title-first
  * ages: 70% min are "18 Years"; 58% max are "Not Specified". Be conservative.

Strategy:
  - conditions:   title-first vocab lookup with stronger blocklist + negation, cap at 2
                  fallback to NER on title only; final fallback to "Not Specified" placeholder
  - study_type:   existing (fine-tuned BERT classifier)
  - sex:          existing
  - ages:         existing strict eligibility regex; conservative defaults when nothing found
  - eligibility:  existing extractors.extract_eligibility (winning strategy)

Usage:
  python pipeline_v6.py --validate --n 50
  python pipeline_v6.py                     # generate submission_v6.csv
"""
import os, sys, re, csv, ast, argparse, time
sys.path.insert(0, os.path.dirname(__file__))
os.environ['LD_LIBRARY_PATH'] = (
    '/usr/local/cuda-12.1/targets/x86_64-linux/lib:'
    + os.environ.get('LD_LIBRARY_PATH', '')
)
os.environ.setdefault('HF_HOME', '/mnt/extra_storage/kkolpetinou/torch_cache/huggingface')

TASK_XLSX = '/home/kkolpetinou/cohort-x-task-1/Task_1.xlsx'
OUTPUT_CSV = '/home/kkolpetinou/cohort-x-task-1/submission_v6.csv'

FIELDNAMES = [
    'pmcids', 'conditions', 'study_type', 'sex',
    'minimum_age', 'maximum_age', 'eligibility_criteria',
]


# ── conditions ───────────────────────────────────────────────────────────────

# Stricter blocklist: only NON-disease junk that consistently leaked through vocab_lookup.
# Keep "cancer", "tumor", "fibrosis" etc — those CAN match GT under BioBERT cosine.
_EXTRA_BLOCKLIST = {
    'staging', 'screening', 'monitoring', 'detection', 'evaluation',
    'learning', 'deep learning', 'machine learning', 'artificial intelligence',
    'fdg pet/ct', 'pet-ct', 'pet/ct', 'pet/mri',  # imaging modalities
    'radiation dosage', 'radiology', 'imaging', 'biopsy',
    'controls', 'healthy', 'volunteers', 'normal',
    'aging', 'birth',
    'observation', 'observational', 'intervention', 'interventional',
    'follow-up', 'follow up',
    'treatment', 'therapy',
    'biomarker', 'biomarkers', 'gene expression',
    'survival', 'mortality',
    'risk', 'risk factors',
}

_NEG_WINDOW = re.compile(
    r'\b(no|without|absent|excluding?|excluded|free of|negative for|denies?|'
    r'absence of|ruled out|cannot have|must not have)\s+(?:\w+\s+){0,8}',
    re.IGNORECASE,
)


def _is_negated(text, term, window=80):
    """Check if `term` appears in negated context within text."""
    text_l = text.lower()
    term_l = term.lower()
    idx = 0
    while True:
        i = text_l.find(term_l, idx)
        if i < 0:
            return False
        # Look at the preceding `window` chars for a negation cue
        start = max(0, i - window)
        before = text_l[start:i]
        for m in _NEG_WINDOW.finditer(before):
            # negation must end close to the term (within ~50 chars)
            if i - (start + m.end()) <= 50:
                idx = i + len(term_l)
                break
        else:
            return False
        if idx == 0:
            return False


_RAG_CACHE = {'index': None, 'embeddings': None}


def _get_rag():
    """Load train index for nearest-train condition borrowing."""
    if _RAG_CACHE['index'] is None:
        import json, numpy as np
        path = os.path.join(os.path.dirname(__file__), 'train_index.json')
        with open(path) as f:
            data = json.load(f)
        embs = np.array(data['embeddings'], dtype='float32')
        norms = np.linalg.norm(embs, axis=1, keepdims=True)
        _RAG_CACHE['embeddings'] = embs / np.maximum(norms, 1e-9)
        _RAG_CACHE['index'] = data['records']
    return _RAG_CACHE['index'], _RAG_CACHE['embeddings']


_PUBMEDBERT = {'tok': None, 'model': None}


def _embed_pubmedbert(text):
    """Embed with the SAME model used to build train_index.json."""
    import torch, numpy as np
    if _PUBMEDBERT['tok'] is None:
        from transformers import AutoTokenizer, AutoModel
        name = 'microsoft/BiomedNLP-PubMedBERT-base-uncased-abstract'
        _PUBMEDBERT['tok'] = AutoTokenizer.from_pretrained(name)
        _PUBMEDBERT['model'] = AutoModel.from_pretrained(name).eval()
    tok, model = _PUBMEDBERT['tok'], _PUBMEDBERT['model']
    enc = tok([text or ''], padding=True, truncation=True, max_length=384, return_tensors='pt')
    with torch.no_grad():
        out = model(**enc)
    mask = enc['attention_mask'].unsqueeze(-1).float()
    emb = (out.last_hidden_state * mask).sum(1) / mask.sum(1).clamp(min=1e-9)
    emb = emb.cpu().numpy()[0]
    return emb / max(np.linalg.norm(emb), 1e-9)


def _rag_conditions(parsed, k=3):
    """Borrow conditions from nearest-K train articles, weighted by similarity."""
    try:
        import numpy as np
        index, embs = _get_rag()
        q = (parsed.get('title', '') or '') + ' ' + (parsed.get('abstract_text', '') or '')[:800]
        q_emb = _embed_pubmedbert(q)
        sims = embs @ q_emb
        top = np.argsort(-sims)[:k]
        cond_counts = {}
        for i in top:
            try:
                conds = ast.literal_eval(index[int(i)]['gt'].get('conditions', '[]') or '[]')
                for c in conds:
                    cond_counts[c] = cond_counts.get(c, 0) + float(sims[int(i)])
            except Exception:
                continue
        if cond_counts:
            ranked = sorted(cond_counts.items(), key=lambda kv: -kv[1])
            return [c for c, _ in ranked[:2]]
    except Exception:
        pass
    return []


def extract_conditions_v6(parsed):
    from extractors import _vocab_lookup

    title = (parsed.get('title', '') or '').strip()
    abstract = (parsed.get('abstract_text', '') or '').strip()
    keywords = parsed.get('keywords', []) or []
    kw_text = ' '.join(keywords)

    def _filter(terms, context):
        out = []
        seen = set()
        for t in terms:
            t_low = t.lower()
            if t_low in _EXTRA_BLOCKLIST:
                continue
            if t_low in seen:
                continue
            if _is_negated(context, t):
                continue
            out.append(t)
            seen.add(t_low)
        return out

    # 1. Strong signal: TITLE vocab match — disease named directly in title
    title_filtered = _filter(_vocab_lookup(title), title)
    if title_filtered:
        title_filtered.sort(key=lambda t: -len(t))
        # If we have a long, specific title hit, use it
        if len(title_filtered[0]) >= 10:
            return str(title_filtered[:2])

    # 2. RAG: borrow from nearest train articles (often nails CT.gov-style condition names)
    rag = _rag_conditions(parsed, k=3)

    # 3. Abstract/keywords vocab as alternative
    abs_filtered = _filter(_vocab_lookup(abstract), abstract)
    kw_filtered = _filter(_vocab_lookup(kw_text), kw_text) if kw_text else []

    # Merge candidates: RAG first (often best), then title (any length), then keywords, then abstract
    candidates = []
    for src in [rag, title_filtered, kw_filtered, abs_filtered]:
        for t in src:
            if t not in candidates:
                candidates.append(t)

    if candidates:
        # Cap at 2 — most GT lists are 1-2 items
        return str(candidates[:2])

    # 4. NER fallback
    try:
        from extractors import _ner_conditions
        ner_hits = _ner_conditions(title) or _ner_conditions(abstract[:1500])
        ner_filtered = _filter(ner_hits, title + ' ' + abstract)
        if ner_filtered:
            ner_filtered.sort(key=lambda t: -len(t))
            return str(ner_filtered[:1])
    except Exception:
        pass

    return str(['Not Specified'])


# ── ages with conservative defaults ──────────────────────────────────────────

def extract_ages_v6(parsed, eligibility_text=''):
    """Strict regex on eligibility section first; conservative defaults otherwise."""
    from llm_extractor_rag import _extract_age_from_article_text, _infer_min_age_from_text

    min_age, max_age = _extract_age_from_article_text(parsed)

    if not min_age:
        # Conservative: infer from "adults" mentions; else default to "18 Years"
        text = eligibility_text + ' ' + parsed.get('abstract_text', '') + ' ' + parsed.get('methods_text', '')
        min_age = _infer_min_age_from_text(eligibility_text, text)
        if not min_age or min_age == 'Not Specified':
            min_age = '18 Years'

    if not max_age:
        # 58% of GT max_age is empty/"Not Specified" — default to that
        max_age = 'Not Specified'

    return min_age, max_age


# ── XLSX I/O ─────────────────────────────────────────────────────────────────

def read_test_pmcids():
    import openpyxl
    wb = openpyxl.load_workbook(TASK_XLSX)
    ws = wb['Test']
    rows = list(ws.iter_rows(values_only=True))
    pmcids = []
    for row in rows[1:]:
        if row[0]:
            pmcids.append(str(int(float(row[0]))).strip())
    return pmcids


def read_train_rows():
    import openpyxl
    wb = openpyxl.load_workbook(TASK_XLSX)
    ws = wb['Train']
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    result = []
    for row in rows[1:]:
        if not row[0]:
            continue
        d = {header[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(header))}
        d['pmcids'] = str(int(float(d['pmcids'])))
        result.append(d)
    return result


# ── Process one article ──────────────────────────────────────────────────────

def process(pmcid):
    from nxml_parser import parse_nxml
    from extractors import extract_sex, extract_study_type, extract_eligibility

    parsed = parse_nxml(pmcid)
    if parsed is None:
        return {
            'pmcids': pmcid,
            'conditions': str(['Not Specified']),
            'study_type': 'INTERVENTIONAL',
            'sex': 'ALL',
            'minimum_age': '18 Years',
            'maximum_age': 'Not Specified',
            'eligibility_criteria': 'Not Specified',
        }

    eligibility = extract_eligibility(parsed)
    # Cap eligibility length — GT max is ~12K, median 700; over 5000 is mostly noise
    if eligibility and len(eligibility) > 5000:
        eligibility = eligibility[:5000].rsplit('* ', 1)[0].rstrip() or eligibility[:5000]
    conditions = extract_conditions_v6(parsed)
    study_type = extract_study_type(parsed)
    sex = extract_sex(parsed)
    min_age, max_age = extract_ages_v6(parsed, eligibility_text=eligibility)

    return {
        'pmcids': pmcid,
        'conditions': conditions,
        'study_type': study_type,
        'sex': sex,
        'minimum_age': min_age,
        'maximum_age': max_age,
        'eligibility_criteria': eligibility,
    }


# ── Validation ───────────────────────────────────────────────────────────────

def validate(n=50, start=0):
    from evaluate import evaluate_row

    train_rows = read_train_rows()
    subset = train_rows[start:start + n]
    print(f"\nValidating articles {start}-{start + len(subset) - 1} ({len(subset)})\n", flush=True)

    FIELDS = ['conditions', 'study_type', 'sex', 'minimum_age', 'maximum_age',
              'eligibility_criteria']
    totals = {f: 0.0 for f in FIELDS}
    results = []

    for i, gt in enumerate(subset, 1):
        pmcid = gt['pmcids']
        t0 = time.time()
        row = process(pmcid)
        elapsed = time.time() - t0
        pred = {f: row.get(f, '') for f in FIELDS}
        gt_fields = {f: gt.get(f, '') for f in FIELDS}
        scores = evaluate_row(pred, gt_fields)
        mean = sum(scores.values()) / len(scores)
        for f in FIELDS:
            totals[f] += scores[f]
        results.append({'pmcid': pmcid, 'scores': scores, 'pred': pred, 'gt': gt_fields})
        print(f"[{i:3d}/{len(subset)}] PMC{pmcid:>8s} ({elapsed:.1f}s)  "
              f"fm3s={scores['eligibility_criteria']:.3f}  "
              f"cond={scores['conditions']:.3f}  "
              f"type={scores['study_type']:.3f}  "
              f"sex={scores['sex']:.2f}  "
              f"min={scores['minimum_age']:.2f}  "
              f"max={scores['maximum_age']:.2f}  "
              f"mean={mean:.3f}", flush=True)

    n_done = len(subset)
    print(f"\n{'=' * 70}")
    print(f"AVERAGES over {n_done} train articles")
    print(f"{'=' * 70}")
    for f in FIELDS:
        print(f"  {f:<25} {totals[f] / n_done:.4f}")
    print(f"  {'MEAN':<25} {sum(totals.values()) / (n_done * len(FIELDS)):.4f}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--validate', action='store_true')
    parser.add_argument('--n', type=int, default=50)
    parser.add_argument('--start', type=int, default=0)
    args = parser.parse_args()

    if args.validate:
        validate(n=args.n, start=args.start)
        return

    pmcids = read_test_pmcids()
    print(f"Processing {len(pmcids)} test articles → {OUTPUT_CSV}", flush=True)
    rows = []
    for i, pmcid in enumerate(pmcids, 1):
        t0 = time.time()
        row = process(pmcid)
        rows.append(row)
        if i % 20 == 0 or i == len(pmcids):
            print(f"  [{i}/{len(pmcids)}] PMC{pmcid} done ({time.time() - t0:.2f}s)", flush=True)

    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)
    print(f"\nWrote {len(rows)} rows → {OUTPUT_CSV}")


if __name__ == '__main__':
    main()
