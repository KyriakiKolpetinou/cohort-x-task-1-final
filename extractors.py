"""
Step 2+3: Rule-based field extractors for CohortX Task 1.
Each extractor takes the dict returned by nxml_parser.parse_nxml()
and returns the value for one field.
"""

import re
import ast
import os

# ── study_type: fine-tuned PubMedBERT ───────────────────────────────────────

_bert_tokenizer = None
_bert_model = None
_BERT_MODEL_DIR = os.path.join(os.path.dirname(__file__), 'models', 'study_type_classifier')

# TF-IDF fallback (used if BERT model not available)
_study_type_clf = None
_study_type_vec = None


def _load_bert_classifier():
    global _bert_tokenizer, _bert_model
    if not os.path.exists(_BERT_MODEL_DIR):
        return False
    try:
        from transformers import AutoTokenizer, AutoModelForSequenceClassification
        import torch
        _bert_tokenizer = AutoTokenizer.from_pretrained(_BERT_MODEL_DIR)
        _bert_model = AutoModelForSequenceClassification.from_pretrained(_BERT_MODEL_DIR)
        _bert_model.eval()
        return True
    except Exception:
        return False


def _train_study_type_classifier():
    global _study_type_clf, _study_type_vec
    import openpyxl
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.linear_model import LogisticRegression

    xlsx = os.path.join(os.path.dirname(__file__), 'Task_1.xlsx')
    if not os.path.exists(xlsx):
        return

    wb = openpyxl.load_workbook(xlsx)
    ws = wb['Train']
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else '' for c in rows[0]]
    pmcid_col = header.index('pmcids')
    type_col  = header.index('study_type')

    texts, labels = [], []
    import sys as _sys
    _sys.path.insert(0, os.path.dirname(__file__))
    from nxml_parser import parse_nxml
    for row in rows[1:]:
        if not row[pmcid_col] or not row[type_col]:
            continue
        label = str(row[type_col]).strip().upper()
        if label not in ('INTERVENTIONAL', 'OBSERVATIONAL'):
            continue
        pmcid = str(int(float(row[pmcid_col])))
        d = parse_nxml(pmcid)
        if d is None:
            continue
        full = (d.get('full_body_text', '') or d.get('abstract_text', '') or '')
        text = (d.get('title', '') + ' ' + full).lower()
        text = re.sub(r'\s+', ' ', text)[:4000]
        texts.append(text)
        labels.append(label)

    if len(texts) < 10:
        return

    vec = TfidfVectorizer(max_features=5000, ngram_range=(1, 2), sublinear_tf=True)
    X = vec.fit_transform(texts)
    clf = LogisticRegression(max_iter=1000, C=1.0, class_weight='balanced')
    clf.fit(X, labels)
    _study_type_vec = vec
    _study_type_clf = clf


def extract_study_type(d):
    """Return 'INTERVENTIONAL' or 'OBSERVATIONAL'.
    Priority: fine-tuned PubMedBERT → TF-IDF+LR → keyword rule-based.
    """
    global _bert_tokenizer, _bert_model, _study_type_clf, _study_type_vec

    # Try BERT first
    if _bert_model is None:
        _load_bert_classifier()

    if _bert_model is not None:
        import torch
        full = (d.get('full_body_text', '') or d.get('abstract_text', '') or '')
        text = (d.get('title', '') + ' ' + full).lower()
        text = re.sub(r'\s+', ' ', text).strip()[:4000]
        inputs = _bert_tokenizer(text, truncation=True, max_length=512, return_tensors='pt')
        with torch.no_grad():
            logits = _bert_model(**inputs).logits
        pred_id = logits.argmax(dim=-1).item()
        return _bert_model.config.id2label[pred_id]

    # Fallback: TF-IDF + LR
    if _study_type_clf is None:
        _train_study_type_classifier()
    if _study_type_clf is not None:
        full = (d.get('full_body_text', '') or d.get('abstract_text', '') or '')
        text = (d.get('title', '') + ' ' + full).lower()
        text = re.sub(r'\s+', ' ', text)[:4000]
        X = _study_type_vec.transform([text])
        return _study_type_clf.predict(X)[0]

    return _extract_study_type_rule_based(d)

# ── Helpers ─────────────────────────────────────────────────────────────────

def _combined_text(d, sections=('eligibility_text', 'methods_text', 'abstract_text')):
    """Concatenate multiple text sources from a parsed article dict."""
    parts = [d.get(s, '') for s in sections]
    return ' '.join(p for p in parts if p)


def _title_abstract(d):
    return (d.get('title', '') + ' ' + d.get('abstract_text', '')).lower()


# ── study_type ───────────────────────────────────────────────────────────────

# Each tuple: (keyword, score)  positive = INTERVENTIONAL, negative = OBSERVATIONAL
_STUDY_TYPE_SCORES = [
    # Strong INTERVENTIONAL signals
    (r'\brandomized\b',           +4),
    (r'\brandomised\b',           +4),
    (r'\brandomization\b',        +3),
    (r'\brandomisation\b',        +3),
    (r'\bdouble.blind\b',         +4),
    (r'\bplacebo.controlled\b',   +4),
    (r'\bplacebo\b',              +3),
    (r'\bcontrolled trial\b',     +4),
    (r'\bclinical trial\b',       +2),
    (r'\bphase (?:i|ii|iii|iv)\b',+3),
    (r'\bphase [1-4]\b',          +3),
    (r'\btreatment arm\b',        +3),
    (r'\bintervention arm\b',     +3),
    (r'\bsingle.blind\b',         +2),
    (r'\bcrossover\b',            +2),
    (r'\binterventional\b',       +2),
    (r'\brct\b',                  +2),
    # Weaker INTERVENTIONAL
    (r'\bprotocol\b',             +1),
    (r'\btreatment group\b',      +1),
    # Strong OBSERVATIONAL signals
    (r'\bretrospective\b',        -4),
    (r'\bobservational study\b',  -4),
    (r'\bobservational design\b', -4),
    (r'\bcohort study\b',         -3),
    (r'\bcase.control\b',         -3),
    (r'\bcross.sectional\b',      -2),
    (r'\bmeta.analysis\b',        -3),
    (r'\bsystematic review\b',    -2),
    (r'\bregistry.based\b',       -3),
    (r'\bepidemiological study\b',-2),
    (r'\bprospective cohort\b',   -2),
    (r'\bcase series\b',          -2),
    # Weaker OBSERVATIONAL
    (r'\bobservational\b',        -1),
    (r'\bcohort\b',               -1),
    (r'\bregistry\b',             -1),
]

_STUDY_TYPE_PATTERNS = [(re.compile(p, re.IGNORECASE), s) for p, s in _STUDY_TYPE_SCORES]


_NCT_PAT = re.compile(r'\bNCT\d{8}\b')

# Patterns for study design — ONLY applied to abstract to avoid noise from cited studies
_OBS_ABSTRACT = re.compile(
    r'\b(?:retrospective|observational\s+study|observational\s+cohort|'
    r'cross.sectional\s+study|case.control\s+study|registry.based|'
    r'prospective\s+observational|prospective\s+cohort\s+study)\b',
    re.IGNORECASE)
_INTERV_ABSTRACT = re.compile(
    r'\b(?:randomized?\s+controlled\s+trial|randomised?\s+controlled\s+trial|'
    r'double.blind(?:ed)?|placebo.controlled|phase\s+(?:i|ii|iii|iv)\s+(?:study|trial)|'
    r'single.blind(?:ed)?|open.label\s+(?:trial|study))\b',
    re.IGNORECASE)


def _extract_study_type_rule_based(d):
    """Return 'INTERVENTIONAL' or 'OBSERVATIONAL'.
    Uses title + abstract for keyword scoring. Also applies precise
    abstract-specific patterns and NCT presence as extra signals.
    Kept as fallback when the TF-IDF classifier is unavailable.
    """
    abstract = d.get('abstract_text', '')
    title = d.get('title', '')
    text = title + ' ' + abstract

    score = 0
    for pat, s in _STUDY_TYPE_PATTERNS:
        if pat.search(text):
            score += s

    # Precise abstract-only patterns (reduce noise vs full-body)
    if _OBS_ABSTRACT.search(abstract):
        score -= 3
    if _INTERV_ABSTRACT.search(abstract):
        score += 3

    # NCT registration number in body → weak INTERVENTIONAL signal
    full = d.get('all_paragraphs_text', '') + ' ' + abstract
    if _NCT_PAT.search(full):
        score += 1

    # Default: INTERVENTIONAL (majority class, 63.5%)
    return 'OBSERVATIONAL' if score < 0 else 'INTERVENTIONAL'


# ── sex ─────────────────────────────────────────────────────────────────────

# Keywords that strongly imply MALE-only study
# Kept tight: only unambiguous male-exclusive disease/anatomy terms
_MALE_PATTERNS = [
    r'\bprostate\s+cancer\b', r'\bprostate\s+carcinoma\b', r'\bprostatectomy\b',
    r'\bprostate.specific\s+antigen\b', r'\bpsa\s+level\b',
    r'\btesticular\s+cancer\b', r'\btesticular\s+torsion\b',
    r'\bpenile\b', r'\berectile\s+dysfunction\b',
    r'\bmale[s]?\s+only\b', r'\bmen\s+only\b',
    r'\bonly\s+men\b', r'\bonly\s+male\b',
    r'\ball\s+(?:patients\s+)?were\s+male\b',
]
# Kept tight: require multi-word phrases to avoid matching "cervical spine",
# "breast imaging in all patients", etc.
_FEMALE_PATTERNS = [
    r'\bbreast\s+cancer\s+patient',
    r'\bbreast\s+cancer\s+(?:cohort|study|trial|population|subject)',
    r'\bHER2.positive\b', r'\bHER2\+\b', r'\btrastuzumab\b', r'\bpertuzumab\b',
    r'\bovarian\s+cancer\b', r'\bovarian\s+carcinoma\b',
    r'\bcervical\s+cancer\b', r'\bcervical\s+carcinoma\b',
    r'\buterine\s+cancer\b', r'\bendometrial\s+cancer\b', r'\bendometrial\s+carcinoma\b',
    r'\bvulvar\s+cancer\b', r'\bvaginal\s+cancer\b',
    r'\bgynecologic(?:al)?\s+(?:cancer|malignancy|oncology)\b',
    r'\bgynaecologic(?:al)?\s+(?:cancer|malignancy|oncology)\b',
    r'\bendometriosis\b',
    r'\bplacenta[l]?\b', r'\bplacental\s+(?:function|imaging|MRI)\b',
    r'\bmaternal.fetal\b', r'\bpregnancy\s+(?:MRI|imaging|complications)\b',
    r'\bnipple\s+discharge\b', r'\bmammograph\b',
    r'\bpelvic\s+floor\s+(?:disorder|dysfunction|defect)\b',
    r'\bfemale[s]?\s+only\b', r'\bwomen\s+only\b',
    r'\bonly\s+women\b', r'\bonly\s+female\b',
    r'\ball\s+(?:patients\s+)?were\s+female\b',
]

_MALE_PAT   = [re.compile(p, re.IGNORECASE) for p in _MALE_PATTERNS]
_FEMALE_PAT = [re.compile(p, re.IGNORECASE) for p in _FEMALE_PATTERNS]


def extract_sex(d):
    """Return 'MALE', 'FEMALE', or 'ALL'."""
    # Search over title + abstract + methods + eligibility
    text = _combined_text(d, ('title', 'abstract_text', 'methods_text', 'eligibility_text'))

    male_hits   = sum(1 for p in _MALE_PAT   if p.search(text))
    female_hits = sum(1 for p in _FEMALE_PAT if p.search(text))

    if male_hits > 0 and female_hits == 0:
        return 'MALE'
    if female_hits > 0 and male_hits == 0:
        return 'FEMALE'
    # Both signals present or neither → default ALL
    return 'ALL'


# ── age extraction ───────────────────────────────────────────────────────────

# Patterns that capture a lower bound and/or upper bound age
# Group 1: lower age  Group 2: upper age  (either can be None)
_AGE_PATTERNS = [
    # "aged X to/- Y years"  or  "aged between X and Y years"
    (re.compile(
        r'aged?\s+(?:between\s+)?(\d+)\s*(?:to|[-–]|and)\s*(\d+)\s*years?',
        re.IGNORECASE), 'range'),
    # "age ≥ X and ≤ Y years"  (two separate constraints in one clause)
    (re.compile(
        r'age[ds]?\s*(?:≥|>=)\s*(\d+)\s*(?:years?)?\s*(?:and|,)\s*(?:≤|<=)\s*(\d+)\s*years?',
        re.IGNORECASE), 'range'),
    # "X to Y years of age"
    (re.compile(
        r'(\d+)\s*(?:to|[-–])\s*(\d+)\s*years?\s*(?:of\s*age|old)',
        re.IGNORECASE), 'range'),
    # "between X and Y years old"
    (re.compile(
        r'between\s+(\d+)\s+and\s+(\d+)\s*years?',
        re.IGNORECASE), 'range'),
    # "≥ X years" / "> X years" / "older than X years"
    (re.compile(
        r'(?:≥|>=|>\s*=?|older\s+than|at\s+least)\s*(\d+)\s*years?',
        re.IGNORECASE), 'min'),
    # "X years or older / over / above"
    (re.compile(
        r'(\d+)\s*years?\s*(?:of\s*age\s*)?(?:or\s*(?:older|over|above)|and\s+(?:older|over|above))',
        re.IGNORECASE), 'min'),
    # "minimum age X" / "aged ≥ X"
    (re.compile(
        r'(?:minimum\s+age|aged?\s*(?:≥|>=))\s*(\d+)',
        re.IGNORECASE), 'min'),
    # "≤ Y years" / "< Y years" / "younger than Y" / "up to Y years"
    (re.compile(
        r'(?:≤|<=|<\s*=?|younger\s+than|up\s+to|no\s+more\s+than|at\s+most)\s*(\d+)\s*years?',
        re.IGNORECASE), 'max'),
    # "Y years or younger / below"
    (re.compile(
        r'(\d+)\s*years?\s*(?:of\s*age\s*)?(?:or\s*(?:younger|below)|and\s+(?:younger|below))',
        re.IGNORECASE), 'max'),
    # "maximum age Y"
    (re.compile(
        r'(?:maximum\s+age)\s*(\d+)',
        re.IGNORECASE), 'max'),
]


def _parse_ages(text):
    """
    Return (min_age_int_or_None, max_age_int_or_None) from text.
    Takes the most common min and the most common max across all matches.
    """
    from collections import Counter
    min_candidates = Counter()
    max_candidates = Counter()

    for pat, kind in _AGE_PATTERNS:
        for m in pat.finditer(text):
            g = [x for x in m.groups() if x is not None]
            if kind == 'range' and len(g) >= 2:
                lo, hi = int(g[0]), int(g[1])
                if 0 < lo < hi <= 120:
                    min_candidates[lo] += 1
                    max_candidates[hi] += 1
            elif kind == 'min' and g:
                v = int(g[0])
                if 0 < v <= 100:
                    min_candidates[v] += 1
            elif kind == 'max' and g:
                v = int(g[0])
                if 0 < v <= 120:
                    max_candidates[v] += 1

    min_age = min_candidates.most_common(1)[0][0] if min_candidates else None
    max_age = max_candidates.most_common(1)[0][0] if max_candidates else None
    return min_age, max_age


def extract_ages(d):
    """Return (min_age_str, max_age_str) using 'X Years' format or 'Not Specified'.
    Only searches eligibility/methods sections to avoid picking up cohort
    ages from result sections (e.g. 'patients aged 22–86').
    """
    # Search only eligibility and methods sections — avoids results-section noise
    search_text = (d.get('eligibility_text', '') + ' ' +
                   d.get('methods_text', '') + ' ' +
                   d.get('abstract_text', ''))
    min_age, max_age = _parse_ages(search_text)

    # Format
    # Default min_age: "18 Years" — 70% of GT values; empty string for remaining 6%
    # Default max_age: "" (empty) — 58% of GT rows have no max age
    min_str = f"{min_age} Years" if min_age is not None else "18 Years"
    max_str = f"{max_age} Years" if max_age is not None else ""
    return min_str, max_str


# ── eligibility_criteria ─────────────────────────────────────────────────────

_INCL_HEADING = re.compile(
    r'inclusion\s+criteria[:\s]', re.IGNORECASE)
_EXCL_HEADING = re.compile(
    r'exclusion\s+criteria[:\s]', re.IGNORECASE)
_ELIGIBLE_SECTION_KWS = re.compile(
    r'\b(?:inclusion|exclusion|eligib|criteria)\b', re.IGNORECASE)


def _sentences_with_criteria(text):
    """Split text into sentences and return those mentioning inclusion/exclusion."""
    sentences = re.split(r'(?<=[.!?])\s+', text)
    return [s.strip() for s in sentences
            if _ELIGIBLE_SECTION_KWS.search(s) and len(s) > 20]


def _format_bullet_list(items):
    """Format a list of strings as '* item' bullet points."""
    return '\n'.join(f'* {item.strip()}' for item in items if item.strip())


def extract_eligibility(d):
    """
    Return eligibility_criteria string in format:
    'Inclusion Criteria:\n\n* ...\n\nExclusion Criteria:\n\n* ...'

    Strategy (in priority order):
    1. Dedicated section headings (inclusion/exclusion section titles)
    2. Inline heading inside any section or paragraph ('Inclusion criteria: ...')
    3. Inline prose signals ('patients were enrolled if', 'we excluded patients who')
    4. Relevant methods paragraph fallback
    """
    incl_parts = []
    excl_parts = []

    # ── Pass 1: dedicated section headings ──────────────────────────────────
    for sec in d.get('body_sections', []):
        title_lower = sec['title'].lower()
        text = sec['text']
        if not text:
            continue
        if 'inclusion' in title_lower and 'exclusion' not in title_lower:
            incl_parts.append(text)
        elif 'exclusion' in title_lower and 'inclusion' not in title_lower:
            excl_parts.append(text)
        elif any(kw in title_lower for kw in ['eligib', 'criteria', 'participants',
                                               'patient', 'subject', 'enroll']):
            # May contain inline inclusion/exclusion headings
            m_incl = _INCL_HEADING.search(text)
            m_excl = _EXCL_HEADING.search(text)
            if m_incl and m_excl and m_incl.start() < m_excl.start():
                incl_parts.append(text[m_incl.end():m_excl.start()].strip())
                excl_parts.append(text[m_excl.end():].strip())
            elif m_incl:
                incl_parts.append(text[m_incl.end():].strip())
            elif m_excl:
                excl_parts.append(text[m_excl.end():].strip())
            elif not incl_parts:
                incl_parts.append(text)

    # ── Pass 2: scan all paragraphs for inline criteria headings ────────────
    if not incl_parts and not excl_parts:
        all_text = d.get('all_paragraphs_text', '') or d.get('full_body_text', '')
        m_incl = _INCL_HEADING.search(all_text)
        m_excl = _EXCL_HEADING.search(all_text)
        if m_incl and m_excl and m_incl.start() < m_excl.start():
            incl_raw = all_text[m_incl.end():m_excl.start()].strip()
            excl_raw = all_text[m_excl.end():m_excl.end()+1500].strip()
            incl_parts.append(incl_raw)
            excl_parts.append(excl_raw)
        elif m_incl:
            incl_parts.append(all_text[m_incl.end():m_incl.end()+1500].strip())
        elif m_excl:
            excl_parts.append(all_text[m_excl.end():m_excl.end()+1500].strip())

    # ── Pass 3: prose-signal paragraphs ─────────────────────────────────────
    _INCL_PROSE = re.compile(
        r'(?:patient|participant|subject|individual)s?\s+'
        r'(?:were\s+)?(?:eligible|included?|enrolled?|recruited?)\s+if\b|'
        r'we\s+(?:enrolled?|included?|recruited?)\s+(?:patient|participant|adult|subject)',
        re.IGNORECASE)
    _EXCL_PROSE = re.compile(
        r'(?:patient|participant|subject|individual)s?\s+'
        r'(?:were\s+)?excluded?\s+(?:if|from|who)\b|'
        r'we\s+excluded?\s+(?:patient|participant|subject)',
        re.IGNORECASE)

    if not incl_parts:
        for para in d.get('all_paragraphs', []):
            if _INCL_PROSE.search(para):
                incl_parts.append(para[:800])
                break
    if not excl_parts:
        for para in d.get('all_paragraphs', []):
            if _EXCL_PROSE.search(para):
                excl_parts.append(para[:800])
                break

    # ── Pass 4: fallback — first methods paragraph with any criteria signal ──
    if not incl_parts and not excl_parts:
        methods = d.get('methods_text', '') or d.get('full_body_text', '')
        if methods:
            incl_parts.append(methods[:600])

    # ── Format output ────────────────────────────────────────────────────────
    def clean_items(parts):
        raw = ' '.join(parts)
        # Split on numbered lists, bullets, semicolons after sentences
        items = re.split(r'\s*[\n•·]\s*|\s*\(\s*[ivxIVX\d]+\s*\)\s*|\s*\d+\)\s+', raw)
        items = [i.strip().rstrip(';.,') for i in items if len(i.strip()) > 12]
        return items[:15]  # cap at 15 bullets

    incl_items = clean_items(incl_parts)
    excl_items = clean_items(excl_parts)

    output_parts = []
    if incl_items:
        output_parts.append('Inclusion Criteria:\n\n' + _format_bullet_list(incl_items))
    if excl_items:
        output_parts.append('Exclusion Criteria:\n\n' + _format_bullet_list(excl_items))

    return '\n\n'.join(output_parts) if output_parts else 'Not Specified'


# ── conditions ───────────────────────────────────────────────────────────────

# Single words that are never diseases on their own
_GENERIC_SINGLE_WORDS = {
    'patients', 'patient', 'subjects', 'subject', 'participants', 'individuals',
    'adults', 'children', 'healthy', 'normal', 'controls', 'volunteers',
    'diagnosis', 'prognosis', 'outcome', 'outcomes', 'survival', 'mortality',
    'treatment', 'therapy', 'treatments', 'therapies', 'surgery', 'procedure',
    'imaging', 'scan', 'mri', 'ct', 'pet', 'spect', 'ultrasound',
    'study', 'trial', 'cohort', 'analysis', 'evaluation', 'assessment',
    'sensitivity', 'specificity', 'accuracy', 'efficacy', 'safety',
    'protein', 'gene', 'marker', 'biomarker', 'dose', 'drug',
    'lung', 'brain', 'liver', 'kidney', 'heart', 'bone',  # anatomy alone
    'hospital', 'clinic', 'university', 'department',
    'background', 'introduction', 'conclusion', 'method', 'result',
    'detection', 'classification', 'segmentation', 'quantification',
    'marking', 'staging', 'grading', 'scoring', 'monitoring',
}

_NON_DISEASE_PAT = re.compile(
    r'\b(?:magnetic\s+resonance(?:\s+imaging)?|computed\s+tomography|'
    r'positron\s+emission|fluorodeoxyglucose|indocyanine\s+green|gadolinium|'
    r'contrast[\s\-](?:enhanced|agent)|artificial\s+intelligence|'
    r'machine\s+learning|deep\s+learning|neural\s+network|'
    r'radiomics|radiomic|algorithm|classifier|model|'
    r'sensitivity|specificity|hazard\s+ratio|odds\s+ratio|'
    r'radiation\s+(?:dose|therapy|treatment)|injection\s+protocol|'
    r'follow.up|follow\s+up|data\s+collection|image\s+(?:acquisition|processing)|'
    r'surgical\s+(?:resection|procedure|technique)|'
    r'(?:intravenous|oral|subcutaneous)\s+(?:injection|administration))\b',
    re.IGNORECASE)

# Lazy-load scispaCy so importing extractors is fast even without spacy
_nlp = None

def _get_nlp():
    global _nlp
    if _nlp is None:
        import spacy
        _nlp = spacy.load('en_core_sci_sm')
    return _nlp


def _build_cond_vocab():
    """Load GT conditions from Task_1.xlsx and build lookup structures."""
    xlsx = '/home/kkolpetinou/cohort-x-task-1/Task_1.xlsx'
    if not os.path.exists(xlsx):
        return {}, []
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    ws = wb['Train']
    vocab = {}   # lowercase → canonical form
    for row in list(ws.iter_rows(values_only=True))[1:]:
        if row[1]:
            try:
                for c in ast.literal_eval(row[1]):
                    c = c.strip()
                    vocab[c.lower()] = c
            except Exception:
                pass
    # Sort by length desc so longer matches win over shorter
    sorted_conds = sorted(vocab.keys(), key=len, reverse=True)
    return vocab, sorted_conds


# These terms appear in GT vocab but are too generic to be useful as conditions
_VOCAB_BLOCKLIST = {
    'patients', 'patient', 'healthy', 'healthy volunteers', 'normal',
    'controls', 'lung', 'heart', 'brain', 'liver', 'kidney', 'bone',
    'treatment', 'therapy', 'diagnosis', 'surgery', 'procedure',
    'imaging', 'mri', 'ct', 'pet', 'scan', 'ultrasound',
    'study', 'trial', 'cohort', 'analysis',
    'adults', 'children', 'volunteers', 'subjects',
    'artificial intelligence', 'machine learning', 'deep learning',
    'magnetic resonance imaging', 'computed tomography',
    'positron emission tomography', 'positron-emission tomography',
    'sentinel lymph node', 'lymph node',
    'radiation', 'radiotherapy', 'chemotherapy', 'immunotherapy',
}

_COND_VOCAB, _COND_LIST = _build_cond_vocab()
# Remove blocklisted terms from the lookup list
_COND_LIST = [c for c in _COND_LIST if c not in _VOCAB_BLOCKLIST]


def _vocab_lookup(text):
    """Find GT conditions that appear in text (case-insensitive)."""
    text_lower = text.lower()
    found = []
    seen = set()
    for c_lower in _COND_LIST:
        canonical = _COND_VOCAB[c_lower]
        if canonical in seen:
            continue
        # Require the condition to appear as a phrase (with word boundaries)
        pat = re.compile(r'\b' + re.escape(c_lower) + r'\b', re.IGNORECASE)
        if pat.search(text_lower):
            found.append(canonical)
            seen.add(canonical)
    return found


_DISEASE_INDICATORS = re.compile(
    r'\b(?:cancer|carcinoma|tumor|tumour|sarcoma|lymphoma|leukemia|leukaemia|'
    r'melanoma|glioma|glioblastoma|adenoma|myeloma|'
    r'disease|disorder|syndrome|condition|'
    r'infection|infect(?:ion|ious)|sepsis|'
    r'failure|insufficiency|deficiency|'
    r'stroke|embolism|thrombosis|infarct(?:ion)?|haemorrhage|hemorrhage|'
    r'inflammation|fibrosis|sclerosis|stenosis|'
    r'diabetes|hypertension|hypotension|'
    r'pneumonia|bronchitis|asthma|copd|'
    r'arthritis|osteoporosis|spondylitis|'
    r'dementia|alzheimer|parkinson|epilepsy|'
    r'malignancy|neoplasm|metastasis|metastatic|'
    r'injury|trauma|fracture|'
    r'COVID|SARS|HIV|AIDS|ARDS|hepatitis)\b',
    re.IGNORECASE)


def _ner_conditions(text):
    """Run scispaCy NER and return disease-like entities.
    Only keeps entities that look like diseases (contain disease indicators
    or are in the GT vocabulary).
    """
    nlp = _get_nlp()
    doc = nlp(text[:2000])
    entities = []
    seen = set()
    for ent in doc.ents:
        t = ent.text.strip()
        t_lower = t.lower()
        words = t_lower.split()
        if len(t) < 3 or len(words) > 6:
            continue
        if t_lower in seen:
            continue
        if len(words) == 1 and t_lower in _GENERIC_SINGLE_WORDS:
            continue
        if _NON_DISEASE_PAT.search(t):
            continue
        if all(w in _GENERIC_SINGLE_WORDS for w in words):
            continue
        # Only keep entity if it has a disease indicator word OR is in GT vocab
        if _DISEASE_INDICATORS.search(t) or t_lower in _COND_VOCAB:
            entities.append(t)
            seen.add(t_lower)
    return entities


def extract_conditions(d):
    """
    Return conditions as a Python list string: "['Disease A', 'Disease B']"

    Strategy: vocabulary lookup only (GT-condition names that appear verbatim).
    Benchmarked: vocab-only F1=20.0% vs combined-with-NER F1=15.1% — NER adds noise.
    """
    search_text = d.get('title', '') + ' ' + d.get('abstract_text', '')

    # Vocabulary lookup — GT condition names appearing in text
    results = _vocab_lookup(search_text)

    # Cap at 8 conditions
    results = results[:8]

    # Guarantee at least one entry
    if not results:
        title_words = [w for w in d.get('title', '').split()
                       if len(w) > 4 and w[0].isupper()]
        results = [title_words[0]] if title_words else ['Not Specified']

    return str(results)
